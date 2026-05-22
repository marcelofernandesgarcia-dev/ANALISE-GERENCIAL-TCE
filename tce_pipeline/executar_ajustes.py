# -*- coding: utf-8 -*-
"""
Script de automação ("Skill") para a ANALISE GERENCIAL TCE.
Este script realiza a limpeza das colunas H, I e J das planilhas,
processa as estatísticas e compila o dashboard interativo.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os
import re
import shutil
import json
import openpyxl
from collections import Counter


# Configurações de caminhos
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHA_DIR = os.path.join(BASE_DIR, "planilha")
BACKUP_DIR = os.path.join(PLANILHA_DIR, "backups")
TEMPLATE_PATH = os.path.join(BASE_DIR, "dashboard_template.html")
OUTPUT_LOCAL_PATH = os.path.join(BASE_DIR, "dashboard.html")
OUTPUT_DOWNLOADS_PATH = r"C:\Users\Mtur.gov\Downloads\dashboard.html"

def clean_text(val):
    if val is None:
        return None
    val_str = str(val)
    # Substitui espaços não quebráveis por espaço comum
    cleaned = val_str.replace('\xa0', ' ')
    # Reduz múltiplos espaços a um único espaço
    cleaned = re.sub(r'\s+', ' ', cleaned)
    # Remove espaços nas pontas e converte em maiúsculas
    return cleaned.strip().upper()

def get_spreadsheets():
    """Retorna uma lista de arquivos de planilha válidos na pasta planilha/"""
    if not os.path.exists(PLANILHA_DIR):
        os.makedirs(PLANILHA_DIR)
        return []
    
    files = []
    for f in os.listdir(PLANILHA_DIR):
        f_path = os.path.join(PLANILHA_DIR, f)
        if os.path.isfile(f_path) and not f.startswith("~$") and f.lower().endswith(('.xlsx', '.xltx')):
            files.append(f)
    return files

def clean_file(filename):
    """Limpa as colunas de Situação Atual, Observação e Encaminhamento Necessário no arquivo"""
    filepath = os.path.join(PLANILHA_DIR, filename)
    print(f"\n[1/3] Limpando arquivo: {filename}")
    
    # Criar pasta de backups se não existir
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
        
    # Backup
    name, ext = os.path.splitext(filename)
    backup_filename = f"{name}_BACKUP{ext}"
    backup_filepath = os.path.join(BACKUP_DIR, backup_filename)
    
    try:
        shutil.copy2(filepath, backup_filepath)
        print(f"  Backup criado em: planilha/backups/{backup_filename}")
    except Exception as e:
        print(f"  Aviso: Não foi possível criar o backup: {e}")
        
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False) # Mantém as fórmulas
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            # Lê a primeira linha (cabeçalhos)
            headers = [cell.value for cell in sheet[1]]
            
            # Detecta as colunas pelo nome (1-based para openpyxl)
            col_h_idx = None
            col_i_idx = None
            col_j_idx = None
            
            for idx, header in enumerate(headers):
                if header:
                    header_upper = str(header).upper()
                    if 'SITUAÇÃO ATUAL' in header_upper:
                        col_h_idx = idx + 1
                    elif 'OBSERVAÇÃO' in header_upper:
                        col_i_idx = idx + 1
                    elif 'ENCAMINHAMENTO NECESSÁRIO' in header_upper:
                        col_j_idx = idx + 1
            
            # Fallbacks baseados na estrutura original se não encontrar por nome
            if col_h_idx is None and len(headers) >= 8:
                col_h_idx = 8
            if col_i_idx is None and len(headers) >= 9:
                col_i_idx = 9
            if col_j_idx is None and len(headers) >= 10:
                col_j_idx = 10
                
            target_cols = [col for col in [col_h_idx, col_i_idx, col_j_idx] if col is not None]
            
            if target_cols:
                changed_count = 0
                total_processed = 0
                for row_idx in range(2, sheet.max_row + 1):
                    for col_idx in target_cols:
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        val = cell.value
                        if val is not None and isinstance(val, str):
                            cleaned = clean_text(val)
                            if cleaned != val:
                                cell.value = cleaned
                                changed_count += 1
                            total_processed += 1
                print(f"  Aba '{sheetname}': processados {total_processed} valores, limpos {changed_count} textos.")
            else:
                print(f"  Aba '{sheetname}': colunas de destino não encontradas.")
                
        print(f"  Salvando alterações em {filename}...")
        wb.save(filepath)
        print(f"  Concluído com sucesso!")
        return True
    except PermissionError:
        print(f"  ERRO: Permissão negada ao salvar {filename}. Certifique-se de fechar a planilha no Excel antes de rodar o script!")
        return False
    except Exception as e:
        print(f"  ERRO inesperado ao processar {filename}: {e}")
        return False

def generate_stats(filename):
    """Lê a planilha limpa e gera as estatísticas em formato JSON"""
    filepath = os.path.join(PLANILHA_DIR, filename)
    print(f"\n[2/3] Extraindo estatísticas de: {filename}")
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        # Seleciona a aba 'Sheet1' ou a ativa por padrão
        if 'Sheet1' in wb.sheetnames:
            sheet = wb['Sheet1']
        else:
            sheet = wb.active
            print(f"  Aviso: Aba 'Sheet1' não encontrada. Usando aba ativa: '{sheet.title}'")
            
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter) # Pula cabeçalho
        
        # Índices dinâmicos baseados no cabeçalho
        col_a_idx = 0  # Nº Transferência
        col_d_idx = 3  # Situação na Plataforma Transferegov
        col_h_idx = 7  # SITUAÇÃO ATUAL
        col_i_idx = 8  # OBSERVAÇÃO
        col_j_idx = 9  # ENCAMINHAMENTO NECESSÁRIO
        
        for idx, header in enumerate(headers):
            if header:
                header_upper = str(header).upper()
                if 'Nº' in header_upper and 'TRANSFER' in header_upper:
                    col_a_idx = idx
                elif 'SITUAÇÃO' in header_upper and 'PLATAFORMA' in header_upper:
                    col_d_idx = idx
                elif 'SITUAÇÃO ATUAL' in header_upper:
                    col_h_idx = idx
                elif 'OBSERVAÇÃO' in header_upper:
                    col_i_idx = idx
                elif 'ENCAMINHAMENTO' in header_upper:
                    col_j_idx = idx

        total_rows = 0
        col_a_vals = []
        col_d_vals = []
        col_h_vals = []
        col_i_vals = []
        col_j_vals = []
        col_d_empty_h = []
        
        for row in rows_iter:
            total_rows += 1
            # Coluna A
            if len(row) > col_a_idx and row[col_a_idx] is not None:
                col_a_vals.append(row[col_a_idx])
            # Coluna D
            if len(row) > col_d_idx and row[col_d_idx] is not None:
                col_d_vals.append(row[col_d_idx])
                
            # Verifica se Coluna H está vazia
            is_h_empty = (len(row) <= col_h_idx or 
                          row[col_h_idx] is None or 
                          str(row[col_h_idx]).strip() == "")
            
            # Coluna H
            if not is_h_empty:
                col_h_vals.append(row[col_h_idx])
            else:
                col_h_vals.append("NÃO INFORMADO")
                # Rastreia coluna D para linhas com H vazia
                if len(row) > col_d_idx and row[col_d_idx] is not None:
                    col_d_empty_h.append(row[col_d_idx])
                else:
                    col_d_empty_h.append("NÃO INFORMADO")
            
            # Coluna I
            if len(row) > col_i_idx and row[col_i_idx] is not None:
                col_i_vals.append(row[col_i_idx])
            else:
                col_i_vals.append("SEM OBSERVAÇÃO")
                
            # Coluna J
            if len(row) > col_j_idx and row[col_j_idx] is not None:
                col_j_vals.append(row[col_j_idx])
            else:
                col_j_vals.append("SEM ENCAMINHAMENTO")

        total_a = len(col_a_vals)
        total_h_filled = sum(1 for x in col_h_vals if x != "NÃO INFORMADO")
        total_h_empty = sum(1 for x in col_h_vals if x == "NÃO INFORMADO")
        
        # Estatísticas da coluna H
        counter_h = Counter([x for x in col_h_vals if x != "NÃO INFORMADO"])
        h_stats = []
        for val, count in counter_h.most_common():
            percentage = (count / total_h_filled) * 100 if total_h_filled > 0 else 0
            h_stats.append({
                "status": val,
                "count": count,
                "percentage": round(percentage, 2)
            })
            
        # Estatísticas da coluna D (Top 10)
        counter_d = Counter(col_d_vals)
        d_stats = []
        for val, count in counter_d.most_common(10):
            d_stats.append({
                "status_plataforma": val,
                "count": count,
                "percentage": round((count / total_rows) * 100, 2) if total_rows > 0 else 0
            })
            
        # Estatísticas da coluna D para registros com H vazia
        counter_d_empty_h = Counter(col_d_empty_h)
        d_empty_h_stats = []
        for val, count in counter_d_empty_h.most_common():
            percentage = (count / total_h_empty) * 100 if total_h_empty > 0 else 0
            d_empty_h_stats.append({
                "status_plataforma": val,
                "count": count,
                "percentage": round(percentage, 2)
            })
            
        # Estatísticas da coluna I (Top 15)
        counter_i = Counter([x for x in col_i_vals if x != "SEM OBSERVAÇÃO"])
        i_stats = []
        for val, count in counter_i.most_common(15):
            i_stats.append({
                "observacao": val,
                "count": count
            })
            
        # Estatísticas da coluna J (Top 15)
        counter_j = Counter([x for x in col_j_vals if x != "SEM ENCAMINHAMENTO" and x != "NÃO HÁ"])
        j_stats = []
        for val, count in counter_j.most_common(15):
            j_stats.append({
                "encaminhamento": val,
                "count": count
            })
            
        data_dict = {
            "total_a": total_a,
            "total_h_filled": total_h_filled,
            "total_h_empty": total_h_empty,
            "h_stats": h_stats,
            "d_stats": d_stats,
            "d_empty_h_stats": d_empty_h_stats,
            "i_stats": i_stats,
            "j_stats": j_stats
        }
        
        # Salva dados locais para depuração/histórico
        debug_json_path = os.path.join(BASE_DIR, "dashboard_data.json")
        with open(debug_json_path, "w", encoding="utf-8") as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=2)
            
        print(f"  Estatísticas processadas. Total A: {total_a}")
        return data_dict
    except Exception as e:
        print(f"  ERRO ao gerar estatísticas para {filename}: {e}")
        return None

def compile_dashboard(data_dict):
    """Substitui o JSON com estatísticas no template do dashboard e salva nos destinos"""
    print(f"\n[3/3] Compilando dashboard...")
    if not os.path.exists(TEMPLATE_PATH):
        print(f"  ERRO: Template {TEMPLATE_PATH} não encontrado!")
        return False
        
    try:
        json_data_str = json.dumps(data_dict, ensure_ascii=False)
        
        with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
            html_template = f.read()
            
        compiled_html = html_template.replace("/*JSON_DATA_PLACEHOLDER*/", json_data_str)
        
        # Salva localmente na pasta do projeto
        with open(OUTPUT_LOCAL_PATH, "w", encoding="utf-8") as f:
            f.write(compiled_html)
        print(f"  Dashboard salvo localmente em: {OUTPUT_LOCAL_PATH}")
        
        # Salva na pasta de Downloads (onde o usuário acessa)
        downloads_dir = os.path.dirname(OUTPUT_DOWNLOADS_PATH)
        if os.path.exists(downloads_dir):
            with open(OUTPUT_DOWNLOADS_PATH, "w", encoding="utf-8") as f:
                f.write(compiled_html)
            print(f"  Dashboard atualizado com sucesso em Downloads: {OUTPUT_DOWNLOADS_PATH}")
        else:
            print(f"  Aviso: Pasta Downloads '{downloads_dir}' não encontrada. Não foi possível exportar para lá.")
            
        print("\n=== PIPELINE CONCLUÍDO COM SUCESSO! ===")
        return True
    except Exception as e:
        print(f"  ERRO ao compilar o dashboard: {e}")
        return False

def main():
    print("===========================================")
    print("  INICIANDO PIPELINE DE AJUSTES E DASHBOARD")
    print("===========================================")
    
    spreadsheets = get_spreadsheets()
    if not spreadsheets:
        print("\nERRO: Nenhuma planilha (.xlsx ou .xltx) encontrada na pasta 'planilha'.")
        print("Por favor, cole a planilha que deseja processar nessa pasta.")
        return
        
    print(f"Planilhas encontradas para limpeza: {spreadsheets}")
    
    # 1. Limpa todas as planilhas encontradas
    for f in spreadsheets:
        clean_file(f)
        
    # 2. Identifica qual será a planilha principal para o dashboard
    # Preferência para a planilha padrão; caso contrário, a de maior tamanho.
    main_file = None
    default_name = "Painel Transferegov_planiha 2008 a 2023.xlsx"
    if default_name in spreadsheets:
        main_file = default_name
    else:
        # Pega a planilha de maior tamanho (desprezando os backups que ficam na pasta backups/)
        xlsx_files = [f for f in spreadsheets if f.lower().endswith('.xlsx')]
        if xlsx_files:
            main_file = max(xlsx_files, key=lambda f: os.path.getsize(os.path.join(PLANILHA_DIR, f)))
        else:
            main_file = spreadsheets[0] # Fallback
            
    print(f"\nPlanilha escolhida como fonte do Dashboard: {main_file}")
    
    # 3. Gera estatísticas
    stats = generate_stats(main_file)
    if stats is None:
        print("\nFalha ao extrair as estatísticas.")
        return
        
    # 4. Compila o Dashboard
    compile_dashboard(stats)

if __name__ == "__main__":
    main()
